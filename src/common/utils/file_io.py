"""File import/export utilities for CSV, TSV, and XLSX formats.

Provides helpers for:

* Serialising row dicts to CSV, TSV, and XLSX byte strings
  (:func:`rows_to_csv`, :func:`rows_to_tsv`, :func:`rows_to_xlsx`).
* Parsing uploaded file bytes back into normalised row dicts
  (:func:`parse_file_upload`).
* Validating parsed rows against Pydantic schemas and collecting errors
  (:func:`validate_rows_and_import`).

Used by the IDV bulk-upload endpoints in ``idv/api/v1/``.
"""

from __future__ import annotations

import csv
import io
from typing import Any

import openpyxl
from pydantic import ValidationError

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_BOOL_TRUE = frozenset(("1", "true", "yes", "on"))
_BOOL_FALSE = frozenset(("0", "false", "no", "off"))
_INT_KEYS = frozenset(("id", "make_id", "model_id"))


def rows_to_csv(rows: list[dict], fieldnames: list[str]) -> str:
    """Serialise rows to a CSV string (RFC 4180).

    Args:
        rows: List of flat dicts; extra keys not in ``fieldnames`` are ignored.
        fieldnames: Ordered list of column headers to include.

    Returns:
        A UTF-8 CSV string with CRLF line endings and a header row.
    """
    buf = io.StringIO()
    writer = csv.DictWriter(
        buf, fieldnames=fieldnames, extrasaction="ignore", lineterminator="\r\n"
    )
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue()


def rows_to_tsv(rows: list[dict], fieldnames: list[str]) -> str:
    """Serialise rows to a TSV string (tab-delimited, RFC 4180 variant).

    Args:
        rows: List of flat dicts; extra keys not in ``fieldnames`` are ignored.
        fieldnames: Ordered list of column headers to include.

    Returns:
        A UTF-8 TSV string with CRLF line endings and a header row.
    """
    buf = io.StringIO()
    writer = csv.DictWriter(
        buf,
        fieldnames=fieldnames,
        extrasaction="ignore",
        delimiter="\t",
        lineterminator="\r\n",
    )
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue()


def rows_to_xlsx(
    rows: list[dict], fieldnames: list[str], sheet_name: str = "Sheet1"
) -> bytes:
    """Serialise rows to an XLSX workbook and return its raw bytes.

    Args:
        rows: List of flat dicts; extra keys not in ``fieldnames`` are ignored.
        fieldnames: Ordered list of column headers (first worksheet row).
        sheet_name: Title of the active worksheet. Defaults to ``"Sheet1"``.

    Returns:
        Raw bytes of a valid XLSX file suitable for streaming to the client.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(fieldnames)
    for row in rows:
        ws.append([str(row.get(f, "")) for f in fieldnames])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _normalize_file_row(row: dict) -> dict[str, Any]:
    """Coerce raw string/cell values from a flat file to Python types.

    * ``is_active`` → bool
    * ``make_id`` / ``model_id`` / ``id`` → int
    * Strips whitespace from string values; drops ``None``-keyed entries.

    Args:
        row: A raw dict produced by a CSV/TSV/XLSX reader.

    Returns:
        A new dict with values coerced to appropriate Python types.
    """
    result: dict[str, Any] = {}
    for key, value in row.items():
        if key is None:
            continue
        key = str(key).strip()
        if value is None or (isinstance(value, str) and value.strip() == ""):
            continue
        if isinstance(value, str):
            value = value.strip()

        if key == "is_active":
            if isinstance(value, bool):
                result[key] = value
            elif isinstance(value, (int, float)):
                result[key] = bool(value)
            else:
                result[key] = str(value).lower() in _BOOL_TRUE
        elif key in _INT_KEYS:
            try:
                result[key] = int(value)
            except (ValueError, TypeError):
                result[key] = value  # let Pydantic produce the error message
        else:
            result[key] = value
    return result


def _parse_csv_bytes(content: bytes) -> list[dict]:
    """Parse UTF-8 (or UTF-8-BOM) CSV bytes into a list of raw dicts."""
    reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
    return list(reader)


def _parse_tsv_bytes(content: bytes) -> list[dict]:
    """Parse UTF-8 (or UTF-8-BOM) TSV bytes into a list of raw dicts."""
    reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")), delimiter="\t")
    return list(reader)


def _parse_xlsx_bytes(content: bytes) -> list[dict]:
    """Parse XLSX bytes into a list of raw dicts using the first sheet."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    return [dict(zip(headers, row, strict=False)) for row in rows[1:]]


def parse_file_upload(content: bytes, filename: str) -> list[dict[str, Any]]:
    """Detect file format from *filename* extension and parse to normalised rows.

    Supported extensions: ``.csv``, ``.tsv``, ``.xlsx``.  Each row is
    normalised via :func:`_normalize_file_row` before being returned.

    Args:
        content: Raw bytes of the uploaded file.
        filename: Original filename used to detect the format by extension.

    Returns:
        List of normalised row dicts ready for Pydantic validation.

    Raises:
        ValueError: If the file extension is not one of ``.csv``, ``.tsv``,
            or ``.xlsx``.
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext == "csv":
        raw = _parse_csv_bytes(content)
    elif ext == "tsv":
        raw = _parse_tsv_bytes(content)
    elif ext == "xlsx":
        raw = _parse_xlsx_bytes(content)
    else:
        raise ValueError(
            f"Unsupported file extension '.{ext}'. Supported formats: .csv, .tsv, .xlsx"
        )
    return [_normalize_file_row(r) for r in raw]


def validate_rows_and_import(rows, item_schema_cls, bulk_schema_cls):
    """Validate *rows* through *item_schema_cls* and return a (schema, errors) tuple.

    Invalid rows are collected in *parse_errors* without aborting the batch.
    Valid rows are packed into a ``bulk_schema_cls`` instance.

    Args:
        rows: Iterable of raw row dicts (output of :func:`parse_file_upload`).
        item_schema_cls: Pydantic model class for validating a single row.
        bulk_schema_cls: Pydantic model class wrapping a list of validated items;
            must accept an ``items`` keyword argument.

    Returns:
        A 2-tuple ``(bulk_schema | None, parse_errors)`` where
        ``bulk_schema`` is ``None`` when no rows passed validation and
        ``parse_errors`` is a list of dicts with ``index`` and ``detail``
        keys for each failed row.
    """
    valid_items = []
    parse_errors: list[dict] = []
    for idx, row in enumerate(rows):
        try:
            valid_items.append(item_schema_cls.model_validate(row))
        except ValidationError as exc:
            parse_errors.append({"index": idx, "detail": exc.errors()})
    schema = bulk_schema_cls(items=valid_items) if valid_items else None
    return schema, parse_errors
